import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

wb = openpyxl.Workbook()

# ===================================================
# カラーパレット
# ===================================================
C_BG_DARK   = "1A1A2E"   # シート背景相当（セル塗り）
C_BG_HEADER = "16213E"   # ヘッダー背景
C_BG_INPUT  = "0F3460"   # 入力セル背景
C_BG_CALC   = "1A1A2E"   # 計算セル背景
C_BG_TOTAL  = "E94560"   # 合計行背景
C_BG_CASH   = "0D7377"   # キャッシュ残高行背景
C_BG_SUBSIDY= "533483"   # 補助金行背景
C_BG_WARN   = "C84B31"   # 警告背景
C_BG_OK     = "2D6A4F"   # 安全背景
C_BG_SECTION= "222831"   # セクション背景

C_TEXT_WHITE  = "FFFFFF"
C_TEXT_YELLOW = "FFD700"
C_TEXT_GREEN  = "00FF88"
C_TEXT_RED    = "FF6B6B"
C_TEXT_LIGHT  = "B0BEC5"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=C_TEXT_WHITE, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Meiryo UI")

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(top=True, bottom=True, left=True, right=True):
    s = Side(style="thin", color="444444")
    n = Side(style=None)
    return Border(
        top=s if top else n,
        bottom=s if bottom else n,
        left=s if left else n,
        right=s if right else n,
    )

def money_fmt(ws, cell):
    ws[cell].number_format = '#,##0'

def pct_fmt(ws, cell):
    ws[cell].number_format = '0%'

# ===================================================
# シート1: 入力パネル（シナリオB推奨）
# ===================================================
ws = wb.active
ws.title = "CFシミュレーション"
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 90

# 列幅設定
col_widths = {
    "A": 4,   # インデント
    "B": 32,  # 項目名
    "C": 16,  # Year 1
    "D": 16,  # Year 2
    "E": 16,  # Year 3
    "F": 4,   # スペーサー
    "G": 20,  # 備考
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# 行高さ
for row in range(1, 120):
    ws.row_dimensions[row].height = 20

# ===================================================
# タイトル
# ===================================================
ws.merge_cells("A1:G1")
ws["A1"] = "Figuraffiti  3ヵ年 CFシミュレーション（自己資金300万円）"
ws["A1"].fill = fill(C_BG_HEADER)
ws["A1"].font = font(bold=True, size=14, color=C_TEXT_YELLOW)
ws["A1"].alignment = align("center")
ws.row_dimensions[1].height = 34

ws.merge_cells("A2:G2")
ws["A2"] = "■ 黄色セルに数値を入力してください。収支・残高は自動計算されます。"
ws["A2"].fill = fill("2C2C54")
ws["A2"].font = font(color="FFD700", size=10, italic=True)
ws["A2"].alignment = align("left")

# ヘッダー行
def write_header(row):
    headers = {"B": "項目", "C": "Year 1", "D": "Year 2", "E": "Year 3", "G": "備考"}
    for col, label in headers.items():
        c = ws[f"{col}{row}"]
        c.value = label
        c.fill = fill(C_BG_HEADER)
        c.font = font(bold=True, size=10)
        c.alignment = align("center")
        c.border = thin_border()

write_header(3)
ws.row_dimensions[3].height = 24

# ===================================================
# ヘルパー: セクションタイトル行
# ===================================================
def section_title(row, label, color=C_BG_SECTION):
    ws.merge_cells(f"A{row}:G{row}")
    c = ws[f"A{row}"]
    c.value = f"  {label}"
    c.fill = fill(color)
    c.font = font(bold=True, size=10, color=C_TEXT_YELLOW)
    c.alignment = align("left")
    c.border = thin_border()
    ws.row_dimensions[row].height = 22

# ===================================================
# ヘルパー: 入力行（黄色セル）
# ===================================================
def input_row(row, label, y1, y2, y3, note="", indent=True):
    prefix = "    " if indent else ""
    ws[f"B{row}"] = prefix + label
    ws[f"B{row}"].fill = fill(C_BG_CALC)
    ws[f"B{row}"].font = font(size=10)
    ws[f"B{row}"].alignment = align("left")
    ws[f"B{row}"].border = thin_border()

    for col, val in zip(["C", "D", "E"], [y1, y2, y3]):
        c = ws[f"{col}{row}"]
        c.value = val
        c.fill = fill("1C3A5E")  # 入力セル: 濃いブルー
        c.font = font(color=C_TEXT_YELLOW, bold=True, size=10)
        c.alignment = align("right")
        c.number_format = '#,##0'
        c.border = thin_border()

    ws[f"G{row}"] = note
    ws[f"G{row}"].fill = fill(C_BG_CALC)
    ws[f"G{row}"].font = font(color=C_TEXT_LIGHT, size=9, italic=True)
    ws[f"G{row}"].alignment = align("left", wrap=True)
    ws[f"G{row}"].border = thin_border()

# ===================================================
# ヘルパー: 計算行（SUM数式）
# ===================================================
def sum_row(row, label, start_row, end_row, bg=C_BG_TOTAL, text_color=C_TEXT_WHITE, note=""):
    ws[f"B{row}"] = f"  ▶ {label}"
    ws[f"B{row}"].fill = fill(bg)
    ws[f"B{row}"].font = font(bold=True, size=10, color=text_color)
    ws[f"B{row}"].alignment = align("left")
    ws[f"B{row}"].border = thin_border()

    for col in ["C", "D", "E"]:
        c = ws[f"{col}{row}"]
        c.value = f"=SUM({col}{start_row}:{col}{end_row})"
        c.fill = fill(bg)
        c.font = font(bold=True, color=text_color, size=10)
        c.alignment = align("right")
        c.number_format = '#,##0'
        c.border = thin_border()

    ws[f"G{row}"] = note
    ws[f"G{row}"].fill = fill(bg)
    ws[f"G{row}"].font = font(color=C_TEXT_LIGHT, size=9)
    ws[f"G{row}"].alignment = align("left")
    ws[f"G{row}"].border = thin_border()

# ===================================================
# ヘルパー: 計算行（カスタム数式）
# ===================================================
def formula_row(row, label, f_c, f_d, f_e, bg=C_BG_CALC, text_color=C_TEXT_WHITE, bold=False, note=""):
    ws[f"B{row}"] = label
    ws[f"B{row}"].fill = fill(bg)
    ws[f"B{row}"].font = font(bold=bold, size=10, color=text_color)
    ws[f"B{row}"].alignment = align("left")
    ws[f"B{row}"].border = thin_border()

    for col, f in zip(["C", "D", "E"], [f_c, f_d, f_e]):
        c = ws[f"{col}{row}"]
        c.value = f
        c.fill = fill(bg)
        c.font = font(bold=bold, color=text_color, size=10)
        c.alignment = align("right")
        c.number_format = '#,##0'
        c.border = thin_border()

    ws[f"G{row}"] = note
    ws[f"G{row}"].fill = fill(bg)
    ws[f"G{row}"].font = font(color=C_TEXT_LIGHT, size=9, italic=True)
    ws[f"G{row}"].alignment = align("left", wrap=True)
    ws[f"G{row}"].border = thin_border()

def spacer(row):
    for col in ["A","B","C","D","E","F","G"]:
        c = ws[f"{col}{row}"]
        c.fill = fill("111111")
        c.value = ""
    ws.row_dimensions[row].height = 8

# ===================================================
# データ入力
# ===================================================
r = 4

# ---- 収入の部 ----
section_title(r, "【収入の部】", "0D3B66"); r += 1
IN_START = r

input_row(r, "自己資金",              300,   0,    0,   "スタート時の自己資金"); r += 1
input_row(r, "融資（日本公庫）",      500,   0,    0,   "新事業活動促進資金"); r += 1
input_row(r, "売上（フィギュア販売）", 80,  400,  900,  "EC・イベント・ライセンス含む"); r += 1
input_row(r, "補助金入金",              0,  195,   33,  "前年分の補助金が翌年入金"); r += 1
input_row(r, "助成金入金",              0,  100,  300,  "展示会助成/Buy TOKYO等"); r += 1
IN_END = r - 1

TOTAL_IN_ROW = r
sum_row(r, "総収入", IN_START, IN_END, bg="0D3B66", text_color="00FFCC"); r += 1

spacer(r); r += 1

# ---- 支出の部 ----
section_title(r, "【支出の部】", "4A0E0E"); r += 1
OUT_START = r

# 設備・初期費用
section_title(r, "  ▸ 設備・初期費用", "2A1A1A"); r += 1
input_row(r, "3Dプリンター追加（2台）",  80,  0,   0,  "1台40万×2台 ※補助金対象"); r += 1
input_row(r, "事務所初期費用",            5,  0,   0,  "シェアオフィス入会金等"); r += 1

# 固定費
section_title(r, "  ▸ 固定費", "2A1A1A"); r += 1
input_row(r, "事務所賃料",              60, 60,  120,  "Y1-2:シェアオフィス5万/月 Y3:固定10万/月"); r += 1
input_row(r, "サーバー・クラウド・API", 12, 12,   24,  "月1万→月2万(Y3)"); r += 1
input_row(r, "コンテンツ制作ツール",    18, 18,   18,  "Adobe等 月1.5万"); r += 1
input_row(r, "融資返済（年間）",        60, 60,   60,  "日本公庫 年間返済額"); r += 1

# 制作・開発費
section_title(r, "  ▸ 制作・開発費", "2A1A1A"); r += 1
input_row(r, "フィギュア制作・材料費",  80, 200, 250,  "量産拡大に伴い増加"); r += 1
input_row(r, "アプリ改修・システム費",   0,  60,   0,  "自前実装のため外注費なし"); r += 1
input_row(r, "EC構築・運営費",           0,  50,   0,  "持続化補助金対象"); r += 1

# マーケティング費
section_title(r, "  ▸ マーケティング・販促費", "2A1A1A"); r += 1
input_row(r, "SNS広告費",              100, 120, 100,  "TikTok/Instagram広告"); r += 1
input_row(r, "SNS運用外注費",           90,   0,   0,  "立ち上げ3ヶ月集中 30万×3 ※補助金対象"); r += 1
input_row(r, "インフルエンサー依頼費",  30,  20,  30,  "10万/回×3回(Y1)"); r += 1
input_row(r, "展示会出展費",             0, 150,   0,  "助成金対象"); r += 1
input_row(r, "コラボ・ライセンス管理費", 0,   0, 200,  "Y3からIP展開本格化"); r += 1
input_row(r, "大規模プロモーション費",   0,   0, 200,  "海外展開準備含む"); r += 1

# 人件費
section_title(r, "  ▸ 人件費", "2A1A1A"); r += 1
input_row(r, "人件費（業務委託・パート）", 0, 120, 300, "Y3は正社員採用も検討"); r += 1

# 諸経費
section_title(r, "  ▸ 諸経費", "2A1A1A"); r += 1
input_row(r, "諸経費（交通・材料・専門家）", 60, 70, 80, ""); r += 1

OUT_END = r - 1

TOTAL_OUT_ROW = r
sum_row(r, "総支出", OUT_START, OUT_END, bg="4A0E0E", text_color="FF6B6B"); r += 1

spacer(r); r += 1

# ---- 収支・残高 ----
section_title(r, "【収支・キャッシュ残高】", "0A3D2E"); r += 1

NET_ROW = r
formula_row(r, "  ▶ 単年収支（収入－支出）",
    f"=C{TOTAL_IN_ROW}-C{TOTAL_OUT_ROW}",
    f"=D{TOTAL_IN_ROW}-D{TOTAL_OUT_ROW}",
    f"=E{TOTAL_IN_ROW}-E{TOTAL_OUT_ROW}",
    bg="0A3D2E", text_color="00FF88", bold=True,
    note="プラスなら黒字、マイナスなら赤字"
); r += 1

CASH_ROW = r
formula_row(r, "  ▶ 期末キャッシュ残高",
    f"=300+C{NET_ROW}",
    f"=C{CASH_ROW-1}+D{NET_ROW}",
    f"=D{CASH_ROW-1}+E{NET_ROW}",
    bg=C_BG_CASH, text_color="00FFCC", bold=True,
    note="300万円スタート → 累積残高"
); r += 1

spacer(r); r += 1

# ---- 補助金シミュレーション ----
section_title(r, "【補助金・助成金シミュレーション】", "2D1B69"); r += 1

input_row(r, "補助金対象経費合計",      220,  50,   0,  "3Dプリンター+SNS広告+外注+インフルエンサー"); r += 1
SUB_ELIGIBLE_ROW = r - 1

input_row(r, "補助率",                  0.5, 0.67,  0,  "新事業進出補助金=1/2、持続化=2/3"); r += 1
SUB_RATE_ROW = r - 1

formula_row(r, "  ▶ 補助金交付決定額（翌年入金）",
    f"=INT(C{SUB_ELIGIBLE_ROW}*C{SUB_RATE_ROW})",
    f"=INT(D{SUB_ELIGIBLE_ROW}*D{SUB_RATE_ROW})",
    f"=INT(E{SUB_ELIGIBLE_ROW}*E{SUB_RATE_ROW})",
    bg=C_BG_SUBSIDY, text_color="FFD700", bold=True,
    note="この金額が翌年の「補助金入金」に反映される"
); r += 1
SUB_RESULT_ROW = r - 1

# 補助率セルの数値フォーマット
for col in ["C", "D", "E"]:
    ws[f"{col}{SUB_RATE_ROW}"].number_format = '0%'

spacer(r); r += 1

# ---- 3年間サマリー ----
section_title(r, "【3年間サマリー】", "1A1A2E"); r += 1

formula_row(r, "  3年間累積売上",
    f"=SUM(C6:E6)",
    "", "",
    bg="1A1A2E", text_color="4FC3F7", bold=True,
    note="3年間の売上合計"
); r += 1
# 累積売上は1セルにまとめる
ws[f"C{r-1}"].value = f"=C6+D6+E6"
ws[f"D{r-1}"].value = ""
ws[f"E{r-1}"].value = ""

formula_row(r, "  3年間累積補助金・助成金",
    f"=SUM(C9:E9)+SUM(C10:E10)",
    "", "",
    bg="1A1A2E", text_color=C_TEXT_YELLOW, bold=True,
    note="補助金入金＋助成金入金の合計"
); r += 1
ws[f"C{r-1}"].value = f"=C9+D9+E9+C10+D10+E10"
ws[f"D{r-1}"].value = ""
ws[f"E{r-1}"].value = ""

formula_row(r, "  3年末キャッシュ残高",
    f"=E{CASH_ROW}",
    "", "",
    bg="1A1A2E", text_color="00FF88", bold=True,
    note="最終的な手元資金"
); r += 1
ws[f"C{r-1}"].value = f"=E{CASH_ROW}"
ws[f"D{r-1}"].value = ""
ws[f"E{r-1}"].value = ""

formula_row(r, "  最低キャッシュ残高（3年間）",
    f"=MIN(C{CASH_ROW},D{CASH_ROW},E{CASH_ROW})",
    "", "",
    bg="1A1A2E", text_color="FF6B6B", bold=True,
    note="この値がマイナスなら資金不足が発生"
); r += 1
ws[f"C{r-1}"].value = f"=MIN(C{CASH_ROW},D{CASH_ROW},E{CASH_ROW})"
ws[f"D{r-1}"].value = ""
ws[f"E{r-1}"].value = ""

# ===================================================
# 条件付き書式：単年収支・残高がマイナスなら赤背景
# ===================================================
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill as PF

red_fill   = PF("solid", fgColor="7B2D2D")
green_fill = PF("solid", fgColor="1B4332")

for col in ["C", "D", "E"]:
    # 単年収支
    ws.conditional_formatting.add(
        f"{col}{NET_ROW}",
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill,
                   font=Font(color="FF6B6B", bold=True, name="Meiryo UI"))
    )
    ws.conditional_formatting.add(
        f"{col}{NET_ROW}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=green_fill,
                   font=Font(color="00FF88", bold=True, name="Meiryo UI"))
    )
    # キャッシュ残高
    ws.conditional_formatting.add(
        f"{col}{CASH_ROW}",
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill,
                   font=Font(color="FF6B6B", bold=True, name="Meiryo UI"))
    )
    ws.conditional_formatting.add(
        f"{col}{CASH_ROW}",
        CellIsRule(operator="greaterThanOrEqual", formula=["100"], fill=green_fill,
                   font=Font(color="00FFCC", bold=True, name="Meiryo UI"))
    )

# ===================================================
# 左端の装飾列
# ===================================================
for row in range(1, r + 5):
    c = ws[f"A{row}"]
    c.fill = fill("0D0D1A")

# ===================================================
# シート2: 使い方ガイド
# ===================================================
ws2 = wb.create_sheet("使い方ガイド")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 2
ws2.column_dimensions["B"].width = 60
ws2.column_dimensions["C"].width = 30

guide_items = [
    ("タイトル", "Figuraffiti CFシミュレーション 使い方ガイド"),
    ("", ""),
    ("基本操作", "■ 黄色セルに数値を入力してください"),
    ("", "  → 収支・残高・補助金額が自動計算されます"),
    ("", ""),
    ("入力セルの場所", "■ 収入の部（行4〜）"),
    ("", "  自己資金 / 融資 / 売上 / 補助金入金 / 助成金入金"),
    ("", ""),
    ("", "■ 支出の部"),
    ("", "  設備費 / 固定費 / 制作費 / マーケ費 / 人件費 / 諸経費"),
    ("", ""),
    ("", "■ 補助金シミュレーション"),
    ("", "  補助金対象経費と補助率を変更すると交付決定額が変わります"),
    ("", "  ※ 交付決定額は翌年の「補助金入金」に手動で転記してください"),
    ("", ""),
    ("注意事項", "■ 補助金は後払い（精算払）です"),
    ("", "  → 交付決定年の翌年に入金されます"),
    ("", "  → 交付決定前の支出は対象外です"),
    ("", ""),
    ("", "■ 融資返済は年間60万円（月5万）で設定しています"),
    ("", "  → 実際の返済額に合わせて変更してください"),
    ("", ""),
    ("", "■ キャッシュ残高がマイナスになる場合"),
    ("", "  → 追加融資または支出削減を検討してください"),
    ("", "  → 東京都中小企業制度融資（成長サポート）も活用可能です"),
    ("", ""),
    ("推奨シナリオ", "■ シナリオB（コスト最適化）を推奨"),
    ("", "  事務所: シェアオフィス 月5万円（Year1-2）"),
    ("", "  SNS外注: 立ち上げ3ヶ月のみ（30万×3ヶ月）"),
    ("", "  → 3年間で370万円のキャッシュ改善効果"),
]

for i, (cat, text) in enumerate(guide_items, start=1):
    ws2.row_dimensions[i].height = 22
    if cat == "タイトル":
        ws2.merge_cells(f"B{i}:C{i}")
        c = ws2[f"B{i}"]
        c.value = text
        c.fill = fill(C_BG_HEADER)
        c.font = Font(bold=True, size=14, color=C_TEXT_YELLOW, name="Meiryo UI")
        c.alignment = Alignment(horizontal="center", vertical="center")
    elif cat:
        ws2[f"B{i}"].value = text
        ws2[f"B{i}"].fill = fill("1A2744")
        ws2[f"B{i}"].font = Font(bold=True, size=10, color="FFD700", name="Meiryo UI")
        ws2[f"B{i}"].alignment = Alignment(horizontal="left", vertical="center")
    else:
        ws2[f"B{i}"].value = text
        ws2[f"B{i}"].fill = fill("111827")
        ws2[f"B{i}"].font = Font(size=10, color="B0BEC5", name="Meiryo UI")
        ws2[f"B{i}"].alignment = Alignment(horizontal="left", vertical="center")

    ws2[f"A{i}"].fill = fill("0D0D1A")

# ===================================================
# 保存
# ===================================================
output_path = "/home/ubuntu/figuraffiti_cf/Figuraffiti_CF_Simulation.xlsx"
wb.save(output_path)
print(f"✅ Excelファイルを生成しました: {output_path}")
print(f"   単年収支行: {NET_ROW}行 / キャッシュ残高行: {CASH_ROW}行")
print(f"   補助金交付決定額行: {SUB_RESULT_ROW}行")
